import os
import re
import subprocess
import sys
from pathlib import Path

import pandas as pd
from PySide6.QtCore import QSettings
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ui_GEpmToolUI import Ui_GEpmTool

# ========== 預設参数 ==========
DEFAULT_INITIAL_DIR = Path("/Volumes/SSD 1TB/GEhealthcare")  # or = Path("")


# 弹出窗口选择目标文件或文件夹
def find_path(select_folder=False):
    """使用 PySide6 选择文件或文件夹"""
    if select_folder:
        # 選擇文件夾
        folder_path = QFileDialog.getExistingDirectory(
            None,
            "选择文件夹",
            str(DEFAULT_INITIAL_DIR)
        )
        if not folder_path:
            print("未选择文件夹，操作取消")
            return None
        return folder_path
    else:
        # 選擇文件
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "选择文件",
            str(DEFAULT_INITIAL_DIR),
            "所有文件 (*.*);;Excel 文件 (*.xlsx *.xls)"
        )
        if not file_path:
            print("未选择文件，操作取消")
            return None
        return file_path


def openfolder(path):
    # path = "/Volumes/SSD 1TB/GEhealthcare"
    if sys.platform == "win32":  # Windows
        os.startfile(path)
    elif sys.platform == "darwin":  # macOS
        subprocess.Popen(["open", path])
    else:  # Linux
        subprocess.Popen(["xdg-open", path])


class ExcelProcess:
    def __init__(self, pm_engineer, pm_phone_number, sample_report_path, output_folder, logger=None):
        # 将传入的路径规范为 pathlib.Path（若为空则保留 None）
        self.bessBox = None
        self.output_folder = Path(output_folder) if output_folder else None
        self.sample_report_path = Path(sample_report_path) if sample_report_path else None
        # self.default_report_path = Path("/Volumes/SSD 1TB/GEhealthcare/Doc/report_demo.xlsx")
        self.pm_engineer = pm_engineer
        self.pm_phone_number = pm_phone_number
        self.logger = logger

        # PM规则配置: key为关键字, value为偏移(月数)
        self.pm_rules = {
            "DEFIBRILLATOR": 6,  # 加6个月
            # 可以在此添加其他规则
        }
        self.default_pm_offset = 12  # 默认加12个月

        # demo report 對應需填充的设备数据
        self.col_map = {
            'Asset ID': 2,
            'Location': 3,
            'Remark': 4,
            'Manufacture': 5,
            'Model': 6,
            'Serial No': 7,
            'Description': 8,
            'ZT': 9,
            'HA Work Order No': 10,
            'Service Report Reference': 14,
        }

        # 新增：動態標題匹配，不再使用固定列字母
        # key = 需要的目標欄位名稱, value = 可接受的表頭關鍵字（忽略大小寫）
        self.dynamic_header_rules = {
            'Asset ID': ['asset id'],
            'Hospital': ['hospital'],
            'Location': ['location'],
            'Manufacture': ['manufacture'],
            'Model': ['model'],
            'Serial No': ['serial no'],
            'Description': ['description'],
            'ZT': ['zt'],
            'HA Work Order No': ['ha work order no'],
            'Schedule Date': ['schedule date'],
            'Service Report Reference': ['service report reference'],
            'Caller': ['caller'],
            'Caller Tel': ['caller tel'],
            'Status': ['^status$'] #只接受完全等於「Status」
        }

    def log(self, message: str):
        if self.logger:
            self.logger(message)
        else:
            print(message)

    @staticmethod
    def clean_filename(name):
        if not isinstance(name, str):
            name = str(name)
        # 移除特殊字符，只保留字母、数字、中文、下划线和短横线
        return re.sub(r'[\\/*?:"<>|]', "", name).strip()

    @staticmethod
    def get_output_path(self):
        output_path = MyWindows.get_output_path()
        return output_path

    def total_model(self, processed_df, output_path):
        model_stats = (
            processed_df
            .groupby(['Manufacture', 'Model', 'Description'])
            .size()
            .reset_index(name='Count')
            .sort_values(by=['Description', 'Manufacture', 'Count'], ascending=[True, True, False])
        )
        path = output_path
        total_model_path = path / "TotalModel.xlsx"

        try:
            model_stats.to_excel(total_model_path, index=False)
            # 添加总数量行
            wb = load_workbook(total_model_path)
            ws = wb.active
            total_count = sum(model_stats['Count'])
            ws.append(["", "", "Total", total_count])

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column  # 获取列的数字索引
                column_letter = get_column_letter(column)
                for cell in col:
                    try:
                        if cell.value:
                            lines = str(cell.value).splitlines()
                            cell_length = max(len(line) for line in lines)
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = int(max_length * 1.2) + 6
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(total_model_path)
            self.logger(f"已创建模型统计文件: {total_model_path}")

        except Exception as e:
            self.logger(f"生成模型统计文件失败: {e}")

        return 1

    def generate_location_files(self, location_df, location, output_dir, template_path, chunk_size=20):
        """为特定Location生成分表文件"""
        # 清理Location名称用于文件名
        clean_loc = self.clean_filename(location)
        if not clean_loc:
            self.logger(f"无效的Location名称: {location}")
            return

        # 按Model和Asset ID排序
        sorted_df = location_df.sort_values(
            by=['Model', 'Asset ID'],
            ascending=[True, True]
        ).reset_index(drop=True)

        # 计算需要分成几个文件
        num_chunks = (len(sorted_df) + chunk_size - 1) // chunk_size

        for i in range(num_chunks):
            # 分块处理数据
            start_idx = i * chunk_size
            end_idx = min((i + 1) * chunk_size, len(sorted_df))
            chunk_df = sorted_df.iloc[start_idx:end_idx].copy()

            # 生成文件名
            suffix = f"({chr(65 + i)})" if i > 0 else ""  # A, B, C...
            filename = f"{clean_loc}{suffix}.xlsx"
            output_path = output_dir / filename

            # 创建基于模板的新工作簿
            wb = load_workbook(template_path)
            ws = wb.active
            # 设置打印方向为横向
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            # 解除设备区域的合并单元格（避免写入错误）
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                # 只解除设备数据区域的合并单元格（第8行到第27行）
                if merged_range.min_row >= 6 and merged_range.min_row <= 25:
                    ws.unmerge_cells(str(merged_range))

            for local_idx, (_, row) in enumerate(chunk_df.iterrows()):
                excel_row = local_idx + 6  # local_idx 是每个 chunk 内的行号，从 0 开始

                # 确保行在有效范围内（8-27）
                if excel_row > 25:
                    print(f"警告: 行号 {excel_row} 超出模板范围，跳过")
                    continue

                for field, col in self.col_map.items():
                    value = row.get(field)
                    if pd.notna(value):
                        ws.cell(row=excel_row, column=col).value = value

                # 设置计划日期 (L列)
                schedule_date = row.get('Schedule Date')
                if pd.notna(schedule_date):
                    try:
                        date_obj = pd.to_datetime(schedule_date)
                        formatted_date = date_obj.strftime("%b-%Y")
                        ws.cell(row=excel_row, column=12).value = f"    {formatted_date}"
                        offset = self.default_pm_offset
                        description = row.get('Description')
                        if pd.notna(description):
                            for keyword, months in self.pm_rules.items():
                                if keyword in str(description):
                                    offset = months
                                    break
                        pm_due_date = date_obj + pd.DateOffset(months=offset)
                        ws.cell(row=excel_row, column=11).value = pm_due_date.strftime("%b-%Y")
                    except Exception as e:
                        print(f"日期格式错误: {schedule_date}, 错误: {e}")

            # 设置联系人信息（使用第一条记录的信息）
            if not chunk_df.empty:
                first_row = chunk_df.iloc[0]
                # 设置医院信息（将Hospital内容添加到B4原本文本的末尾）
                if pd.notna(first_row.get('Hospital')):
                    current_value = ws.cell(row=4, column=2).value or ""
                    ws.cell(row=4, column=2).value = f"{current_value}{first_row['Hospital']}"
                if pd.notna(first_row.get('Caller')):
                    ws.cell(row=28, column=5).value = first_row['Caller']  # E30
                if pd.notna(first_row.get('Caller Tel')):
                    ws.cell(row=28, column=7).value = first_row['Caller Tel']  # G30
                if self.pm_engineer:
                    ws.cell(row=27, column=5).value = self.pm_engineer  # E27
                if self.pm_phone_number:
                    ws.cell(row=27, column=7).value = self.pm_phone_number  # G27

            # 保存文件
            wb.save(output_path)
            print(f"已创建分表: {output_path}")

    def preprocess(self):
        """主处理函数"""
        # 获取模板文件路径
        template_path = self.sample_report_path
        if not template_path or not Path(template_path).exists():
            self.logger("错误: 模板文件未找到")
            return 0

        file_path = self.output_folder
        # 获取当前文件的目录
        output_dir = Path(file_path).parent / "Output"
        output_dir.mkdir(parents=True, exist_ok=True)
        self.logger(f"输出目录: {output_dir}")

        try:
            self.logger(f"正在读取总表文件: {file_path}")
            # 读取Excel文件
            df = pd.read_excel(file_path, engine='openpyxl')

        except Exception as e:
            self.logger(f"读取文件失败: {e}")
            return 0

        # 動態搜尋表頭欄位
        resolved_columns = {}
        for target_key, keywords in self.dynamic_header_rules.items():
            found_col = None
            for col in df.columns:
                normalized = str(col).strip().lower()
                if any(re.fullmatch(pattern, normalized) for pattern in keywords):
                    found_col = col
                    break
            resolved_columns[target_key] = found_col
            #print(f"已创建分表: {found_col}") #測試尋找表頭是否正確
        # self.log(f"動態匹配到的欄位: {resolved_columns}")
        
        # 解析 bessList
        bess_assets = []
        if hasattr(self, "bessList") and hasattr(self, "bessBox") and self.bessBox.isChecked():
            text = self.bessList.toPlainText()
            # 去重並保持輸入順序
            raw_assets = [str(a).strip() for a in re.split(r"[\n,]+", text) if a.strip()]
            bess_assets = list(dict.fromkeys(raw_assets))
            self.logger(f"BESS 跟機共: {len(bess_assets)}台")
            
        # 创建新的DataFrame用于处理
        processed_data = []

        # 遍历每一行
        for _, row in df.iterrows():
            row_data = {}

            # 处理每一列
            for col_name, original_col_name in resolved_columns.items():
                if original_col_name is not None:
                    value = row[original_col_name]

                    # 型別處理
                    if col_name in ['Asset ID', 'HA Work Order No']:
                        try:
                            # 尝试转换为整数，如果失败则保持原样
                            value = int(value) if pd.notna(value) else None
                        except:
                            pass

                    row_data[col_name] = value

            # 分别处理Accepted和On Hold，且Location非空且非空字符串
            status_value = row_data.get('Status')
            location_value = row_data.get('Location')
            asset_id = row_data.get('Asset ID')

            # BESS 特殊處理
            asset_id_str = str(asset_id).strip() if pd.notna(asset_id) else ""
            if bess_assets and asset_id_str in bess_assets:
                row_data['__status_group'] = 'BESS'
                row_data['Remark'] = 'BESS'
                processed_data.append(row_data)
                #print(f"已找到bess Asset{row_data}")
                continue  # 已處理，跳過 Accepted/OnHold

            if status_value == 'Accepted' and pd.notna(location_value) and str(location_value).strip() != "":
                row_data['__status_group'] = 'Accepted'
                processed_data.append(row_data)

            elif hasattr(self, "onHoldBox") and self.onHoldBox.isChecked() \
                and status_value == 'On Hold' and pd.notna(location_value) and str(location_value).strip() != "":
                # On Hold 设备：Location 独立分组，并写 Remark
                row_data['__status_group'] = 'On Hold'
                row_data['Remark'] = 'On Hold'
                processed_data.append(row_data)

        # 创建新的DataFrame
        processed_df = pd.DataFrame(processed_data)
        self.logger(f"成功读取文件: 本月共有「 {len(processed_df)} 」部機器")

        if processed_df.empty:
            self.logger("警告: 没有找到有效的Location数据")
            return 0

        # 按Location分组处理（On Hold 和 BESS 设备Location独立分组）
        def group_key(row):
            if row.get('__status_group') == 'On Hold':
                return f"{row['Location']}_OnHold"
            elif row.get('__status_group') == 'BESS':
                return f"{row['Location']}_BESS"
            else:
                return row['Location']

        grouped = processed_df.groupby(processed_df.apply(group_key, axis=1))

        self.logger(f"找到 {len(grouped)} 个不同的Location")

        # 为每个Location生成分表
        for location, group in grouped:
            self.logger(f"处理Location: {location}, 设备数: {len(group)}")
            self.generate_location_files(group, location, output_dir, template_path)

        self.total_model(processed_df, output_dir)

        return 1

    def run(self):
        result = self.preprocess()
        if result:
            self.logger("===========操作完成!==========")
        else:
            self.logger("===========操作失敗!===========")


class MyWindows(QMainWindow, Ui_GEpmTool):
    # 屬性配置
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.bind()
        self.sample_path = None  # config["paths"]["sample_path"]
        self.report_path = None  # config["paths"]["report_path"]
        self.settings = QSettings("GEpmTool", "UserConfig")
        self.load_settings()

    def bind(self):
        self.pushButton.clicked.connect(self.process)
        self.pushButton_2.clicked.connect(self.exit_program)
        self.pushButton_3.clicked.connect(self.open_output_folder)
        self.pushButton_4.clicked.connect(self.log_clear)
        self.toolButton.clicked.connect(self.set_sample_path)
        self.toolButton_2.clicked.connect(self.set_output_path)
        self.bessBox.setChecked(False)
        self.onHoldBox.setChecked(True)
        self.bessList.setPlainText("")

        self.actionGuide.triggered.connect(self.show_guide)

    def load_settings(self):
        self.lineEdit.setText(self.settings.value("pm_engineer", ""))  # 工程師名稱
        self.lineEdit_2.setText(self.settings.value("pm_phone", ""))  # 工程師電話
        self.lineEdit_5.setText(self.settings.value("sample_path", ""))  # 模板路徑
        self.lineEdit_6.setText(self.settings.value("output_path", ""))  # 目標表格路徑
        self.bessBox.setChecked(self.settings.value("bessBox_checked", True, type=bool))
        self.onHoldBox.setChecked(self.settings.value("onHoldBox_checked", True, type=bool))

    def save_settings(self):
        self.settings.setValue("pm_engineer", self.lineEdit.text())
        self.settings.setValue("pm_phone", self.lineEdit_2.text())
        self.settings.setValue("sample_path", self.lineEdit_5.text())
        self.settings.setValue("output_path", self.lineEdit_6.text())
        self.settings.setValue("bessBox_checked", self.bessBox.isChecked())
        self.settings.setValue("onHoldBox_checked", self.onHoldBox.isChecked())

    def set_sample_path(self):
        file_path = find_path(select_folder=False)
        if file_path:
            self.lineEdit_5.setText(file_path)  # 把路徑填寫到 lineEdit_5

    def set_output_path(self):
        file_path = find_path(select_folder=False)
        if file_path:
            self.lineEdit_6.setText(file_path)  # 把路徑填寫到 lineEdit_6

    def path_check(self, line, path_str):
        if os.path.isdir(path_str):  # 如果 lineEdit_6內容為非法路徑
            QMessageBox.warning(self, "提示", f"請選用正確的{line}")
            return 0
        elif not path_str:  # 如果 lineEdit_6內容為空
            QMessageBox.warning(self, "提示", f"{line}不能為空")
            return 0
        else:  # 如果 lineEdit_6 內容為文件
            return 1

    def get_output_path(self):
        if self.path_check("Target File", self.lineEdit_6.text()):
            # 獲取文件跟目錄
            base_dir = os.path.dirname(self.lineEdit_6.text())
            # 在路徑後面拼接 Output 子資料夾
            output_path = os.path.join(base_dir, "Output")
            return output_path
        else:
            return None

    def open_output_folder(self):
        file_path = self.get_output_path()
        if file_path:
            # 打開對應目錄
            openfolder(file_path)

    def process(self):
        pm_engineer = self.lineEdit.text()
        pm_phone_number = self.lineEdit_2.text()
        sample_report_path = self.lineEdit_5.text()
        output_path = self.lineEdit_6.text()
        if self.path_check("Sample_Report_File", sample_report_path):
            if self.path_check("Target File", output_path):
                # 把UI裡對應的信息和接口傳給ExcelProcess類函數運行
                processor = ExcelProcess(
                    pm_engineer,
                    pm_phone_number,
                    sample_report_path,
                    output_path,
                    logger=self.log_output
                    )

                # 將 UI 的 BESS/OnHold 控件和 bessList 傳給 processor
                # 注意：processor 內部會用 hasattr 判斷控件是否存在
                processor.bessBox = getattr(self, "bessBox", None)
                processor.onHoldBox = getattr(self, "onHoldBox", None)
                processor.bessList = getattr(self, "bessList", None)

                processor.run()
                self.save_settings()

    # 退出程序
    def exit_program(self):
        self.close()

    # UI輸出log的接口
    def log_output(self, text):
        """將日誌消息輸出到 UI 的 plainTextEdit"""
        self.plainTextEdit.appendPlainText(text)

    # 清空log的窗口
    def log_clear(self):
        self.plainTextEdit.clear()

    def show_guide(self):
        QMessageBox.information(self, 'Guide',
                                'Step1:填自己名\n'
                                'Step2:填自己電話\n'
                                'Step3:SampleReport->找Koen\n'
                                'Step4:TargetFile->APM獲取PMTaskReport\n'
                                'Step5:按需要勾選OnHold或BESS\n'
                                'Step6:輸入BESS Asset,每個Asset回車換行(如無可跳過)\n'
                                'Step7:點擊Generate\n'
                                'Step8:點擊OutputFolder查看生成文件')


if __name__ == '__main__':
    app = QApplication([])
    window = MyWindows()
    window.show()
    app.exec()
