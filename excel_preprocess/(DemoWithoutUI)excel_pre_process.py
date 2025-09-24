import os
import re
import sys
import tkinter as tk
from pathlib import Path
from tkinter import filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ========== 全局配置参数 ==========
DEFAULT_INITIAL_DIR = Path("/Volumes/SSD 1TB/GEhealthcare") # or = Path("")
DEFAULT_REPORT_PATH = Path("/Volumes/SSD 1TB/GEhealthcare/Doc/report_demo.xlsx")
DEFAULT_OUTPUT_PATH = Path("") #Path("/Volumes/SSD 1TB/GEhealthcare/202508/Output")
DEVICE_HEADER_KEYS = {
    'Asset ID': 'B',
    'Location': 'C',
    'Manufacture': 'E',
    'Model': 'F',
    'Serial No': 'G',
    'Description': 'H',
    'ZT': 'I',
    'HA Work Order No': 'J',
    'Service Report Reference': 'N'
}

# PM规则配置: key为关键字, value为偏移(月数)
PM_RULES = {
    "DEFIBRILLATOR": 6,  # 加6个月
    # 可以在此添加其他规则
}
DEFAULT_PM_OFFSET = 12  # 默认加12个月

"""error check module"""
class ErrCode:
    """标准错误码常量 (符合 0=成功 约定)"""
    SUCCESS = 0  # 操作成功
    INVALID_ARGUMENT = 1  # 无效参数
    FILE_NOT_FOUND = 2  # 文件不存在
    TEMPLATE_NOT_FOUND = 3  # 模板文件未找到
    UNKNOWN_ERROR = 99  # 未知错误


def errcheck(result):
    if result != ErrCode.SUCCESS:
        print(f"err_code:{result}")

#弹出窗口选择目标文件
def find_excel_file():
    # 测试用 - 取消注释以下三行用于测试
    # =================find_excel_file====================================
    file_path = "/Volumes/SSD 1TB/GEhealthcare/2025test/Preventive-Pending Report.xlsx"
    return file_path
    # =====================================================

    # 正式用
    """使用 tkinter 选择文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 创建文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择总表文件",
        filetypes=[
            ("Excel 文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ],
        initialdir=DEFAULT_INITIAL_DIR  # 初始目录
    )
    root.destroy()  # 关闭 tkinter 窗口
    if not file_path:
        print("未选择文件，操作取消")
        return None
    return file_path

#清理文件名中的无效字符
def clean_filename(name):
    if not isinstance(name, str):
        name = str(name)
    # 移除特殊字符，只保留字母、数字、中文、下划线和短横线
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

# 生成全表模型统计文件
def total_model(processed_df,output_dir):
    model_stats = (
        processed_df
        .groupby(['Manufacture','Model','Description'])
        .size()
        .reset_index(name='Count')
        .sort_values(by=['Description', 'Manufacture', 'Count'], ascending=[True, True, False])
    )
    total_model_path = output_dir / "TotalModel.xlsx"
    try:
        model_stats.to_excel(total_model_path, index=False)
        # 添加总数量行
        wb = load_workbook(total_model_path)
        ws = wb.active
        total_count = sum(model_stats['Count'])
        ws.append(["", "", "Total", total_count])

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # 获取列的数字索引
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value:
                        lines = str(cell.value).splitlines()
                        cell_length = max(len(line) for line in lines)
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = int(max_length * 1.2) + 6
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(total_model_path)
        print(f"已创建模型统计文件: {total_model_path}")

    except Exception as e:
        print(f"生成模型统计文件失败: {e}")

    print("处理完成!")
    return ErrCode.SUCCESS

def generate_location_files(location_df, location, output_dir, template_path, chunk_size=20):
    """为特定Location生成分表文件"""
    # 清理Location名称用于文件名
    clean_loc = clean_filename(location)
    if not clean_loc:
        print(f"无效的Location名称: {location}")
        return

    # 按Model和Asset ID排序
    sorted_df = location_df.sort_values(
        by=['Model', 'Asset ID'],
        ascending=[True, True]
    ).reset_index(drop=True)

    # 计算需要分成几个文件
    num_chunks = (len(sorted_df) + chunk_size - 1) // chunk_size

    for i in range(num_chunks):
        # 分块处理数据
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, len(sorted_df))
        chunk_df = sorted_df.iloc[start_idx:end_idx].copy()

        # 生成文件名
        suffix = f"({chr(65 + i)})" if i > 0 else ""  # A, B, C...
        filename = f"{clean_loc}{suffix}.xlsx"
        output_path = output_dir / filename

        # 创建基于模板的新工作簿
        wb = load_workbook(template_path)
        ws = wb.active
        # 设置打印方向为横向
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        # 解除设备区域的合并单元格（避免写入错误）
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # 只解除设备数据区域的合并单元格（第8行到第27行）
            if merged_range.min_row >= 6 and merged_range.min_row <= 25:
                ws.unmerge_cells(str(merged_range))

        # 填充设备数据
        col_map = {
            'Asset ID': 2,
            'Location': 3,
            'Manufacture': 5,
            'Model': 6,
            'Serial No': 7,
            'Description': 8,
            'ZT': 9,
            'HA Work Order No': 10,
            'Service Report Reference': 14,
        }

        for local_idx, (_, row) in enumerate(chunk_df.iterrows()):
            excel_row = local_idx + 6  # local_idx 是每个 chunk 内的行号，从 0 开始

            # 确保行在有效范围内（8-27）
            if excel_row > 25:
                print(f"警告: 行号 {excel_row} 超出模板范围，跳过")
                continue

            for field, col in col_map.items():
                value = row.get(field)
                if pd.notna(value):
                    ws.cell(row=excel_row, column=col).value = value

            # 设置计划日期 (L列)
            schedule_date = row.get('Schedule Date')
            if pd.notna(schedule_date):
                try:
                    date_obj = pd.to_datetime(schedule_date)
                    formatted_date = date_obj.strftime("%b-%Y")
                    ws.cell(row=excel_row, column=12).value = f"    {formatted_date}"
                    offset = DEFAULT_PM_OFFSET
                    description = row.get('Description')
                    if pd.notna(description):
                        for keyword, months in PM_RULES.items():
                            if keyword in str(description):
                                offset = months
                                break
                    pm_due_date = date_obj + pd.DateOffset(months=offset)
                    ws.cell(row=excel_row, column=11).value = pm_due_date.strftime("%b-%Y")
                except Exception as e:
                    print(f"日期格式错误: {schedule_date}, 错误: {e}")

        # 设置联系人信息（使用第一条记录的信息）
        if not chunk_df.empty:
            first_row = chunk_df.iloc[0]
            # 设置医院信息（将Hospital内容添加到B4原本文本的末尾）
            if pd.notna(first_row.get('Hospital')):
                current_value = ws.cell(row=4, column=2).value or ""
                ws.cell(row=4, column=2).value = f"{current_value}{first_row['Hospital']}"
            if pd.notna(first_row.get('Caller')):
                ws.cell(row=28, column=5).value = first_row['Caller']  # E30
            if pd.notna(first_row.get('Caller Tel')):
                ws.cell(row=28, column=7).value = first_row['Caller Tel']  # G30

        # 保存文件
        wb.save(output_path)
        print(f"已创建分表: {output_path}")


def preprocess(file_path):
    """主处理函数"""
    # 获取模板文件路径
    template_path = DEFAULT_REPORT_PATH
    if not template_path or not Path(template_path).exists():
        print("错误: 模板文件未找到")
        return ErrCode.TEMPLATE_NOT_FOUND

    # 获取当前文件的目录
    if not str(DEFAULT_OUTPUT_PATH).strip():
        output_dir = DEFAULT_OUTPUT_PATH
    else:
        output_dir = Path(file_path).parent / "Output"

    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"输出目录: {output_dir}")

    try:
        print(f"正在读取总表文件: {file_path}")
        # 读取Excel文件
        df = pd.read_excel(file_path, engine='openpyxl')
        print(f"成功读取文件: 共 {len(df)} 条记录")

    except Exception as e:
        print(f"读取文件失败: {e}")
        return ErrCode.FILE_NOT_FOUND

    # 列映射定义 (Excel列字母 -> 我们的列名)
    col_mapping = {
        'D': 'Asset ID',
        'F': 'Hospital',
        'K': 'Location',
        'L': 'Manufacture',
        'M': 'Model',
        'N': 'Serial No',
        'O': 'Description',
        'EV': 'ZT',
        'R': 'HA Work Order No',
        'T': 'Schedule Date',
        'U': 'Service Report Reference',
        'I': 'Caller',
        'J': 'Caller Tel'
    }

    # 创建反向映射 (列字母 -> 列索引)
    col_letter_to_index = {}
    for idx, col in enumerate(df.columns):
        col_letter = get_column_letter(idx + 1)
        col_letter_to_index[col_letter] = col

    # 创建新的DataFrame用于处理
    processed_data = []

    # 遍历每一行
    for _, row in df.iterrows():
        row_data = {}

        # 处理每一列
        for col_letter, col_name in col_mapping.items():
            # 检查列字母是否在映射中
            if col_letter in col_letter_to_index:
                original_col_name = col_letter_to_index[col_letter]
                value = row[original_col_name]

                # 处理特定列的数据类型
                if col_name in ['Asset ID', 'HA Work Order No']:
                    try:
                        # 尝试转换为整数，如果失败则保持原样
                        value = int(value) if pd.notna(value) else None
                    except (ValueError, TypeError):
                        pass

                row_data[col_name] = value

        # 只添加有Location数据的行
        if pd.notna(row_data.get('Location')):
            processed_data.append(row_data)

    # 创建新的DataFrame
    processed_df = pd.DataFrame(processed_data)

    if processed_df.empty:
        print("警告: 没有找到有效的Location数据")
        return ErrCode.SUCCESS

    # 按Location分组处理
    grouped = processed_df.groupby('Location')
    print(f"找到 {len(grouped)} 个不同的Location")

    # 为每个Location生成分表
    for location, group in grouped:
        print(f"处理Location: {location}, 设备数: {len(group)}")
        generate_location_files(group, location, output_dir, template_path)

    total_model(processed_df,output_dir)

    print("处理完成!")
    return ErrCode.SUCCESS

def run(target_file):
    print("PM订单分表生成工具")
    print("=" * 60)
    print("功能: 根据总表生成按Location分组的设备分表")
    print("=" * 60)


    if target_file:
        result = preprocess(target_file)
        if result == ErrCode.SUCCESS:
            print("操作成功完成!")
        else:
            print(f"处理完成，但有错误 (代码: {result})")

        # 等待用户输入退出
        if sys.platform == 'win32':
            os.system("pause")
    else:
        print("未选择文件，程序退出")

if __name__ == "__main__":
    run()