import os
import sys
import numpy as np
import inspect
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import filedialog, messagebox
import re

"""error check module"""


class ErrCode:
    """标准错误码常量 (符合 0=成功 约定)"""
    SUCCESS = 0  # 操作成功
    INVALID_ARGUMENT = 1  # 无效参数
    FILE_NOT_FOUND = 2  # 文件不存在
    TEMPLATE_NOT_FOUND = 3  # 模板文件未找到
    UNKNOWN_ERROR = 99  # 未知错误


def errcheck(result):
    frame = inspect.currentframe()
    try:
        # 获取调用栈中上一帧（调用此函数的帧）
        caller_frame = frame.f_back
        line = caller_frame.f_lineno
    finally:
        # 显式删除帧引用以避免内存泄漏
        del frame
    if result != ErrCode.SUCCESS:
        print(f"err_code:{result},line:{line}")


def find_excel_file():
    # 测试用 - 取消注释以下三行用于测试
    # =====================================================
    #file_path = "/Volumes/SSD 1TB/GEhealthcare/202508/Preventive-Pending Report.xlsx"
    #return file_path
    # =====================================================

    # 正式用
    """使用 tkinter 选择文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 创建文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择总表文件",
        filetypes=[
            ("Excel 文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ],
        initialdir="/Volumes/SSD 1TB/GEhealthcare/"  # 初始目录
    )
    root.destroy()  # 关闭 tkinter 窗口
    if not file_path:
        print("未选择文件，操作取消")
        return None
    return file_path


def get_template_path():
    """获取模板文件路径"""
    # 固定模板路径
    template_path = Path("/Volumes/SSD 1TB/GEhealthcare/Doc/report_demo.xlsx")

    if template_path.exists():
        return template_path
    else:
        # 如果固定路径不存在，尝试让用户选择
        root = tk.Tk()
        root.withdraw()
        template_path = filedialog.askopenfilename(
            title="选择模板文件",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialdir="/Volumes/SSD 1TB/GEhealthcare/Doc/"
        )
        root.destroy()
        return template_path if template_path else None


def clean_filename(name):
    """清理文件名中的无效字符"""
    if not isinstance(name, str):
        name = str(name)
    # 移除特殊字符，只保留字母、数字、中文、下划线和短横线
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def total_model(processed_df,output_dir):
    # 生成全表模型统计文件
    model_stats = (
        processed_df
        .groupby(['Model', 'Manufacture'])
        .size()
        .reset_index(name='Count')
        .sort_values(by='Count', ascending=False)
    )
    total_model_path = output_dir / "TotalModel.xlsx"
    try:
        model_stats.to_excel(total_model_path, index=False)
        print(f"已创建模型统计文件: {total_model_path}")
    except Exception as e:
        print(f"生成模型统计文件失败: {e}")

    print("处理完成!")
    return ErrCode.SUCCESS

def generate_location_files(location_df, location, output_dir, template_path, chunk_size=20):
    """为特定Location生成分表文件"""
    # 清理Location名称用于文件名
    clean_loc = clean_filename(location)
    if not clean_loc:
        print(f"无效的Location名称: {location}")
        return

    # 按Model和Asset ID排序
    sorted_df = location_df.sort_values(
        by=['Model', 'Asset ID'],
        ascending=[True, True]
    ).reset_index(drop=True)

    # 计算需要分成几个文件
    num_chunks = (len(sorted_df) + chunk_size - 1) // chunk_size

    for i in range(num_chunks):
        # 分块处理数据
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, len(sorted_df))
        chunk_df = sorted_df.iloc[start_idx:end_idx].copy()

        # 生成文件名
        suffix = f"({chr(65 + i)})" if i > 0 else ""  # A, B, C...
        filename = f"{clean_loc}{suffix}.xlsx"
        output_path = output_dir / filename

        # 创建基于模板的新工作簿
        wb = load_workbook(template_path)
        ws = wb.active

        # 解除设备区域的合并单元格（避免写入错误）
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # 只解除设备数据区域的合并单元格（第2行到第21行）
            if merged_range.min_row >= 2 and merged_range.min_row <= 21:
                ws.unmerge_cells(str(merged_range))

        # 填充设备数据
        for local_idx, (_, row) in enumerate(chunk_df.iterrows()):
            excel_row = local_idx + 2  # local_idx 是每个 chunk 内的行号，从 0 开始


            # 确保行在有效范围内（2-21）
            if excel_row > 21:
                print(f"警告: 行号 {excel_row} 超出模板范围，跳过")
                continue

            # 设置资产ID (B列)
            if pd.notna(row.get('Asset ID')):
                ws.cell(row=excel_row, column=2).value = row['Asset ID']

            # 设置位置 (C列)
            if pd.notna(row.get('Location')):
                ws.cell(row=excel_row, column=3).value = row['Location']

            # 设置制造商 (E列)
            if pd.notna(row.get('Manufacture')):
                ws.cell(row=excel_row, column=5).value = row['Manufacture']

            # 设置型号 (F列)
            if pd.notna(row.get('Model')):
                ws.cell(row=excel_row, column=6).value = row['Model']

            # 设置序列号 (G列)
            if pd.notna(row.get('Serial No')):
                ws.cell(row=excel_row, column=7).value = row['Serial No']

            # 设置描述 (H列)
            if pd.notna(row.get('Description')):
                ws.cell(row=excel_row, column=8).value = row['Description']

            # 设置ZT状态 (I列)
            if pd.notna(row.get('ZT')):
                ws.cell(row=excel_row, column=9).value = row['ZT']

            # 设置HA工单号 (J列)
            if pd.notna(row.get('HA Work Order No')):
                ws.cell(row=excel_row, column=10).value = row['HA Work Order No']

            # 设置服务报告参考 (N列)
            if pd.notna(row.get('Service Report Reference')):
                ws.cell(row=excel_row, column=14).value = row['Service Report Reference']

        # 设置联系人信息（使用第一条记录的信息）
        if not chunk_df.empty:
            first_row = chunk_df.iloc[0]
            if pd.notna(first_row.get('Caller')):
                ws.cell(row=24, column=5).value = first_row['Caller']  # E24
            if pd.notna(first_row.get('Caller Tel')):
                ws.cell(row=24, column=7).value = first_row['Caller Tel']  # G24

        # 保存文件
        wb.save(output_path)
        print(f"已创建分表: {output_path}")


def preprocess(file_path):
    """主处理函数"""
    # 获取模板文件路径
    template_path = get_template_path()
    if not template_path or not Path(template_path).exists():
        print("错误: 模板文件未找到")
        return ErrCode.TEMPLATE_NOT_FOUND

    # 获取当前文件的目录
    output_dir = Path(file_path).parent / 'Output'
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"输出目录: {output_dir}")

    try:
        print(f"正在读取总表文件: {file_path}")
        # 读取Excel文件
        df = pd.read_excel(file_path, engine='openpyxl')
        print(f"成功读取文件: 共 {len(df)} 条记录")

    except Exception as e:
        print(f"读取文件失败: {e}")
        return ErrCode.FILE_NOT_FOUND

    # 列映射定义 (Excel列字母 -> 我们的列名)
    col_mapping = {
        'D': 'Asset ID',
        'K': 'Location',
        'L': 'Manufacture',
        'M': 'Model',
        'N': 'Serial No',
        'O': 'Description',
        'EV': 'ZT',
        'R': 'HA Work Order No',
        'U': 'Service Report Reference',
        'I': 'Caller',
        'J': 'Caller Tel'
    }

    # 创建反向映射 (列字母 -> 列索引)
    col_letter_to_index = {}
    for idx, col in enumerate(df.columns):
        col_letter = get_column_letter(idx + 1)
        col_letter_to_index[col_letter] = col

    # 创建新的DataFrame用于处理
    processed_data = []

    # 遍历每一行
    for _, row in df.iterrows():
        row_data = {}

        # 处理每一列
        for col_letter, col_name in col_mapping.items():
            # 检查列字母是否在映射中
            if col_letter in col_letter_to_index:
                original_col_name = col_letter_to_index[col_letter]
                value = row[original_col_name]

                # 处理特定列的数据类型
                if col_name in ['Asset ID', 'HA Work Order No']:
                    try:
                        # 尝试转换为整数，如果失败则保持原样
                        value = int(value) if pd.notna(value) else None
                    except (ValueError, TypeError):
                        pass

                row_data[col_name] = value

        # 只添加有Location数据的行
        if pd.notna(row_data.get('Location')):
            processed_data.append(row_data)

    # 创建新的DataFrame
    processed_df = pd.DataFrame(processed_data)

    if processed_df.empty:
        print("警告: 没有找到有效的Location数据")
        return ErrCode.SUCCESS

    # 按Location分组处理
    grouped = processed_df.groupby('Location')
    print(f"找到 {len(grouped)} 个不同的Location")

    # 为每个Location生成分表
    for location, group in grouped:
        print(f"处理Location: {location}, 设备数: {len(group)}")
        generate_location_files(group, location, output_dir, template_path)

    total_model(processed_df,output_dir)

    print("处理完成!")
    return ErrCode.SUCCESS


if __name__ == "__main__":
    print("PM订单分表生成工具")
    print("=" * 60)
    print("功能: 根据总表生成按Location分组的设备分表")
    print("=" * 60)

    # 查找Excel文件
    target_file = find_excel_file()
    if target_file:
        result = preprocess(target_file)
        if result == ErrCode.SUCCESS:
            print("操作成功完成!")
        else:
            print(f"处理完成，但有错误 (代码: {result})")

        # 等待用户输入退出
        if sys.platform == 'win32':
            os.system("pause")
    else:
        print("未选择文件，程序退出")